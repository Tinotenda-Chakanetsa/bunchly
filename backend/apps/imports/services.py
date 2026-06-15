"""Data-import business logic (spec §9.14).

The pipeline is two-step: ``validate_batch`` parses a CSV / XLSX upload,
runs every row through the entity-specific validator and writes one
:class:`~apps.imports.models.ImportError` per issue, then ``commit_batch``
re-validates and creates the rows inside a transaction. Templates are
generated from the same column metadata so the source file and the
validator can never drift.
"""
from __future__ import annotations

import csv
import io
from typing import Iterator

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from .entities import EntityDefinition, get_definition
from .enums import ImportStatus
from .models import ImportBatch, ImportError as ImportErrorModel


# --- file parsing ----------------------------------------------------------

def _decode_csv(file_obj) -> Iterator[dict]:
    """Yield row dicts from a CSV upload. Handles a UTF-8 BOM."""
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(raw))
    for row in reader:
        yield {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}


def _decode_xlsx(file_obj) -> Iterator[dict]:
    """Yield row dicts from an XLSX upload. Uses openpyxl (already a dep)."""
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    headers = [str(h).strip() if h is not None else "" for h in header]
    for raw in rows:
        if not any(cell is not None and str(cell).strip() for cell in raw):
            continue
        row = {}
        for i, value in enumerate(raw):
            if i >= len(headers) or not headers[i]:
                continue
            if value is None:
                row[headers[i]] = ""
            elif hasattr(value, "isoformat"):  # datetime / date
                row[headers[i]] = value.isoformat()[:10]
            else:
                row[headers[i]] = str(value).strip()
        yield row


def _parse_upload(filename: str, file_obj) -> list[dict]:
    """Return a list of row dicts from a CSV or XLSX upload."""
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return list(_decode_xlsx(file_obj))
    if name.endswith(".csv"):
        return list(_decode_csv(file_obj))
    raise ValidationError(
        {"file": "Unsupported file type — upload a .csv or .xlsx file."}
    )


# --- templates -------------------------------------------------------------

def template_csv(entity_type: str) -> str:
    """A header-only CSV template with a help-text comment row."""
    definition = get_definition(entity_type)
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(definition.columns)
    writer.writerow(
        [definition.template_help.get(col, "") for col in definition.columns]
    )
    return out.getvalue()


# --- batch lifecycle -------------------------------------------------------

def _record_errors(batch: ImportBatch, errors: list[tuple[int, str, str]]):
    """Persist (row_number, field, error) tuples as :class:`ImportError`s."""
    ImportErrorModel.objects.filter(batch=batch).delete()
    ImportErrorModel.objects.bulk_create([
        ImportErrorModel(
            batch=batch, row_number=row_number, field=field or "", error=message,
        )
        for row_number, field, message in errors
    ])


def _validate_rows(
    tenant, definition: EntityDefinition, rows: list[dict]
) -> tuple[list[tuple[int, str, str]], int]:
    """Apply the entity validator to every row. Returns (errors, valid_count)."""
    errors: list[tuple[int, str, str]] = []
    valid = 0
    for index, row in enumerate(rows, start=2):  # row 1 is the header
        row_errors = definition.validate_row(tenant, row, index)
        if row_errors:
            errors.extend((index, field, message) for field, message in row_errors)
        else:
            valid += 1
    return errors, valid


@transaction.atomic
def validate_batch(
    *, tenant, entity_type: str, filename: str, file_obj, user=None
) -> ImportBatch:
    """Parse + validate an upload, returning a fresh ``validated`` batch."""
    definition = get_definition(entity_type)
    rows = _parse_upload(filename, file_obj)

    batch = ImportBatch.objects.create(
        tenant=tenant,
        entity_type=entity_type,
        original_filename=filename or "",
        total_rows=len(rows),
    )
    errors, valid = _validate_rows(tenant, definition, rows)
    error_row_numbers = {row_number for row_number, _, _ in errors}
    _record_errors(batch, errors)
    batch.valid_rows = valid
    batch.error_rows = len(error_row_numbers)
    batch.status = ImportStatus.VALIDATED
    batch.save(update_fields=[
        "valid_rows", "error_rows", "status", "updated_at",
    ])
    return batch


@transaction.atomic
def commit_batch(
    *, batch: ImportBatch, file_obj, filename: str | None = None, user=None
) -> ImportBatch:
    """Re-validate the upload then create rows for every clean entry."""
    if batch.status == ImportStatus.COMMITTED:
        raise ValidationError("This batch has already been committed.")

    definition = get_definition(batch.entity_type)
    name = filename or batch.original_filename
    rows = _parse_upload(name, file_obj)

    errors, valid = _validate_rows(batch.tenant, definition, rows)
    error_row_numbers = {row_number for row_number, _, _ in errors}
    _record_errors(batch, errors)

    created = 0
    for index, row in enumerate(rows, start=2):
        if index in error_row_numbers:
            continue
        definition.apply_row(batch.tenant, row)
        created += 1

    batch.total_rows = len(rows)
    batch.valid_rows = valid
    batch.error_rows = len(error_row_numbers)
    batch.committed_rows = created
    batch.status = ImportStatus.COMMITTED if created else ImportStatus.FAILED
    batch.committed_at = timezone.now()
    batch.save(update_fields=[
        "total_rows", "valid_rows", "error_rows", "committed_rows",
        "status", "committed_at", "updated_at",
    ])
    return batch
